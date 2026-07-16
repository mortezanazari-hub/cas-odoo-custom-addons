import base64
import csv
import hashlib
import io
import re
from datetime import date, datetime, time, timedelta

import openpyxl

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, ValidationError


def _jalali_to_gregorian(jy, jm, jd):
    jy += 1595
    days = -355668 + 365 * jy + (jy // 33) * 8 + ((jy % 33 + 3) // 4) + jd
    days += (jm - 1) * 31 if jm < 7 else (jm - 7) * 30 + 186
    gy = 400 * (days // 146097); days %= 146097
    if days > 36524:
        days -= 1; gy += 100 * (days // 36524); days %= 36524
        if days >= 365: days += 1
    gy += 4 * (days // 1461); days %= 1461
    if days > 365:
        gy += (days - 1) // 365; days = (days - 1) % 365
    gd = days + 1
    leap = gy % 4 == 0 and (gy % 100 != 0 or gy % 400 == 0)
    months = [31, 29 if leap else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    gm = 1
    for length in months:
        if gd <= length: break
        gd -= length; gm += 1
    return date(gy, gm, gd)


class CasAttendanceImport(models.Model):
    _name = "cas.attendance.import"
    _description = "CAS Audited Attendance Import Batch"
    _inherit = ["mail.thread"]
    _order = "create_date desc, id desc"
    _rec_name = "name"

    name = fields.Char(string="عنوان ورود", required=True, tracking=True)
    company_id = fields.Many2one("res.company", string="شرکت", required=True, default=lambda self: self.env.company, index=True)
    import_type = fields.Selection([
        ("device_punches", "رکورد خام دستگاه"), ("paired_sessions", "فایل ورود و خروج جفت‌شده"),
        ("guard_workbook", "دفتر Excel نگهبانی با شیت هر شخص"),
    ], string="قالب ورودی", required=True, default="device_punches", tracking=True)
    device_id = fields.Many2one("cas.attendance.device", string="دستگاه", ondelete="restrict")
    site_id = fields.Many2one("cas.attendance.site", string="محل", required=True, ondelete="restrict")
    data_file = fields.Binary(string="فایل", required=True, attachment=True)
    filename = fields.Char(string="نام فایل", required=True)
    state = fields.Selection([("draft", "آماده خواندن"), ("review", "در انتظار بازبینی"), ("imported", "واردشده")], string="وضعیت", required=True, default="draft", readonly=True, tracking=True)
    line_ids = fields.One2many("cas.attendance.import.line", "batch_id", string="ردیف‌ها", copy=False)
    total_count = fields.Integer(compute="_compute_counts")
    ready_count = fields.Integer(compute="_compute_counts")
    unmatched_count = fields.Integer(compute="_compute_counts")
    imported_count = fields.Integer(compute="_compute_counts")
    skipped_count = fields.Integer(compute="_compute_counts")
    parsed_at = fields.Datetime(readonly=True)
    imported_at = fields.Datetime(readonly=True)
    imported_by_id = fields.Many2one("res.users", readonly=True)

    @api.depends("line_ids.status")
    def _compute_counts(self):
        for rec in self:
            rec.total_count = len(rec.line_ids)
            rec.ready_count = len(rec.line_ids.filtered(lambda l: l.status == "ready"))
            rec.unmatched_count = len(rec.line_ids.filtered(lambda l: l.status == "unmatched"))
            rec.imported_count = len(rec.line_ids.filtered(lambda l: l.status == "imported"))
            rec.skipped_count = len(rec.line_ids.filtered(lambda l: l.status in {"duplicate", "skipped", "error"}))

    def write(self, vals):
        structural = {"company_id", "import_type", "device_id", "site_id", "data_file", "filename"}
        if structural.intersection(vals) and any(rec.state != "draft" for rec in self):
            raise ValidationError(_("فایل خوانده‌شده قابل جایگزینی نیست؛ یک بچ جدید بسازید."))
        engine = {"state", "parsed_at", "imported_at", "imported_by_id"}
        if engine.intersection(vals) and not self.env.context.get("cas_import_engine"):
            raise AccessError(_("وضعیت ورود فقط از عملیات رسمی تغییر می‌کند."))
        return super().write(vals)

    def unlink(self):
        if any(rec.state != "draft" for rec in self):
            raise ValidationError(_("بچ خوانده‌شده برای ممیزی قابل حذف نیست."))
        return super().unlink()

    @api.model
    def _parse_datetime(self, value, date_value=False):
        if isinstance(value, datetime): return value.replace(second=0, microsecond=0)
        if isinstance(value, date): return datetime.combine(value, time.min)
        if isinstance(value, time) and date_value:
            base = self._parse_date(date_value)
            return datetime.combine(base, value).replace(second=0, microsecond=0) if base else False
        text = str(value or "").strip()
        if not text: return False
        if date_value and re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", text):
            base = self._parse_date(date_value)
            hour, minute = map(int, text.split(":")[:2])
            return datetime.combine(base, time(hour, minute)) if base else False
        match = re.search(r"(1[34]\d{2})[/\-](\d{1,2})[/\-](\d{1,2})(?:\s+(\d{1,2}):(\d{1,2})(?::\d{1,2})?)?", text)
        if match:
            y, m, d = map(int, match.group(1, 2, 3)); base = _jalali_to_gregorian(y, m, d)
            return datetime.combine(base, time(int(match.group(4) or 0), int(match.group(5) or 0)))
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%Y/%m/%d %H:%M"):
            try: return datetime.strptime(text, fmt).replace(second=0, microsecond=0)
            except ValueError: pass
        return False

    @api.model
    def _parse_date(self, value):
        if isinstance(value, datetime): return value.date()
        if isinstance(value, date): return value
        parsed = self._parse_datetime(value)
        return parsed.date() if parsed else False

    def _add_line(self, row_no, key, occurred_at, event_kind, external_suffix, note=False):
        self.ensure_one()
        employee = self.env["cas.attendance.identity"].employee_for("guard" if self.import_type == "guard_workbook" else "device", key, self.company_id)
        uid_seed = f"{self.filename}|{self.import_type}|{row_no}|{key}|{external_suffix}|{occurred_at}"
        external_uid = hashlib.sha256(uid_seed.encode("utf-8")).hexdigest()
        duplicate = self.env["cas.attendance.event"].sudo().search_count([("source", "=", "guard" if self.import_type == "guard_workbook" else "device"), ("external_uid", "=", external_uid), ("company_id", "=", self.company_id.id)])
        return self.env["cas.attendance.import.line"].create({
            "batch_id": self.id, "row_number": row_no, "external_key": str(key or "").strip(),
            "employee_id": employee.id, "occurred_at": occurred_at, "event_kind": event_kind,
            "external_uid": external_uid, "status": "duplicate" if duplicate else ("ready" if employee and occurred_at else "unmatched"),
            "note": note,
        })

    def action_parse(self):
        for batch in self:
            if batch.state != "draft": raise ValidationError(_("این فایل قبلاً خوانده شده است."))
            raw = base64.b64decode(batch.data_file or b"")
            if not raw: raise ValidationError(_("فایل خالی است."))
            if batch.import_type == "guard_workbook": batch._parse_guard_xlsx(raw)
            elif batch.import_type == "paired_sessions": batch._parse_paired(raw)
            else: batch._parse_device(raw)
            if not batch.line_ids: raise ValidationError(_("هیچ ردیف قابل تشخیصی در فایل پیدا نشد."))
            batch.with_context(cas_import_engine=True).write({"state": "review", "parsed_at": fields.Datetime.now()})

    def _workbook(self, raw):
        try: return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc: raise ValidationError(_("فایل Excel قابل خواندن نیست: %s", exc)) from exc

    def _parse_guard_xlsx(self, raw):
        wb = self._workbook(raw)
        for ws in wb.worksheets:
            key = ws.title.strip()
            for row_no, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if len(row) < 10: continue
                day = self._parse_date(row[7])
                if not day: continue
                entry = self._parse_datetime(row[8], row[7])
                exit_ = self._parse_datetime(row[9], row[7])
                if entry: self._add_line(row_no, key, entry, "guard_entry", f"{ws.title}:in")
                if exit_:
                    if entry and exit_ <= entry: exit_ += timedelta(days=1)
                    self._add_line(row_no, key, exit_, "guard_exit", f"{ws.title}:out")

    def _rows_from_file(self, raw):
        if self.filename.lower().endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            return list(csv.DictReader(io.StringIO(text)))
        wb = self._workbook(raw); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return []
        headers = [str(v or "").strip() for v in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def _parse_paired(self, raw):
        for row_no, row in enumerate(self._rows_from_file(raw), start=2):
            key = row.get("employee_id") or row.get("device_id") or row.get("Name") or row.get("EnNo")
            entry = self._parse_datetime(row.get("check_in") or row.get("InTime") or row.get("ورود"))
            exit_ = self._parse_datetime(row.get("check_out") or row.get("OutTime") or row.get("خروج"))
            if entry: self._add_line(row_no, key, entry, "work_start", "in")
            if exit_: self._add_line(row_no, key, exit_, "work_end", "out")

    def _parse_device(self, raw):
        staged = []
        for row_no, row in enumerate(self._rows_from_file(raw), start=2):
            key = row.get("EnNo") or row.get("Name") or row.get("device_id") or row.get("employee_id")
            value = row.get("DateTime") or row.get("Date/Time") or row.get("timestamp") or row.get("تاریخ و زمان")
            occurred = self._parse_datetime(value)
            if key not in (None, "") and occurred: staged.append((row_no, str(key).strip().removesuffix(".0"), occurred))
        grouped = {}
        for item in staged:
            employee = self.env["cas.attendance.identity"].employee_for("device", item[1], self.company_id)
            work_date = self.env["cas.attendance.event"]._resolve_work_date(employee, item[2]) if employee else item[2].date()
            grouped.setdefault((item[1], work_date), []).append(item)
        for (key, _work_date), items in grouped.items():
            items.sort(key=lambda x: x[2])
            self._add_line(items[0][0], key, items[0][2], "work_start", "first")
            if len(items) > 1: self._add_line(items[-1][0], key, items[-1][2], "work_end", "last")
            for row_no, _key, occurred in items[1:-1]:
                self._add_line(row_no, key, occurred, "unknown_entry", "middle", note=_("رکورد میانی؛ برای حالت پیشرفته نیازمند تعیین نوع است."))

    def action_import_ready(self):
        for batch in self:
            if batch.state != "review": raise ValidationError(_("بچ در مرحله بازبینی نیست."))
            if batch.unmatched_count: raise ValidationError(_("ابتدا همه شناسه‌های ناشناس را نگاشت یا ردیف‌ها را رد کنید."))
            ready = batch.line_ids.filtered(lambda l: l.status == "ready")
            for line in ready.sorted(lambda l: (l.occurred_at, l.id)):
                try:
                    event = self.env["cas.attendance.event"].with_context(cas_attendance_supervisor=True).create({
                        "employee_id": line.employee_id.id, "occurred_at": line.occurred_at,
                        "source": "guard" if batch.import_type == "guard_workbook" else "device",
                        "event_kind": line.event_kind, "site_id": batch.site_id.id,
                        "device_id": batch.device_id.id if batch.device_id else False,
                        "external_uid": line.external_uid, "note": _("ورودی بچ %s، ردیف %s", batch.name, line.row_number),
                    })
                    line.with_context(cas_import_engine=True).write({"status": "imported", "event_id": event.id})
                except Exception as exc:
                    line.with_context(cas_import_engine=True).write({"status": "error", "error_message": str(exc)})
            batch.with_context(cas_import_engine=True).write({"state": "imported", "imported_at": fields.Datetime.now(), "imported_by_id": self.env.user.id})


class CasAttendanceImportLine(models.Model):
    _name = "cas.attendance.import.line"
    _description = "CAS Attendance Import Staging Line"
    _order = "batch_id, occurred_at, row_number, id"

    batch_id = fields.Many2one("cas.attendance.import", required=True, ondelete="cascade", index=True)
    company_id = fields.Many2one(related="batch_id.company_id", store=True, index=True)
    row_number = fields.Integer(string="ردیف فایل", readonly=True)
    external_key = fields.Char(string="شناسه خارجی", readonly=True, index=True)
    employee_id = fields.Many2one("hr.employee", string="کارمند", ondelete="restrict", index=True)
    occurred_at = fields.Datetime(string="زمان رخداد", readonly=True, index=True)
    event_kind = fields.Selection(selection=lambda self: self.env["cas.attendance.event"]._fields["event_kind"].selection, string="نوع رخداد")
    external_uid = fields.Char(readonly=True, index=True)
    status = fields.Selection([("ready", "آماده"), ("unmatched", "شناسه ناشناس"), ("duplicate", "تکراری"), ("skipped", "ردشده"), ("imported", "واردشده"), ("error", "خطا")], required=True, readonly=True, index=True)
    event_id = fields.Many2one("cas.attendance.event", string="رخداد ساخته‌شده", readonly=True, ondelete="restrict")
    note = fields.Text(string="یادداشت", readonly=True)
    error_message = fields.Text(string="خطا", readonly=True)

    def write(self, vals):
        if self.env.context.get("cas_import_engine"):
            return super().write(vals)
        allowed = {"employee_id", "event_kind"}
        if set(vals) - allowed or any(line.batch_id.state != "review" or line.status not in {"ready", "unmatched"} for line in self):
            raise AccessError(_("ردیف ورودی فقط در مرحله بازبینی و در فیلدهای مجاز قابل اصلاح است."))
        result = super().write(vals)
        for line in self:
            if line.employee_id and line.occurred_at and line.event_kind and line.status == "unmatched":
                line.with_context(cas_import_engine=True).write({"status": "ready"})
        return result

    def unlink(self):
        if any(line.batch_id.state != "draft" for line in self): raise ValidationError(_("ردیف بچ خوانده‌شده قابل حذف نیست."))
        return super().unlink()

    def action_skip(self):
        self.with_context(cas_import_engine=True).write({"status": "skipped"})
