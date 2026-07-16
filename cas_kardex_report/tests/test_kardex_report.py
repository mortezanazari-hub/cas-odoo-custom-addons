import io
from datetime import date

import openpyxl

from odoo.exceptions import ValidationError
from odoo.tests.common import TransactionCase, tagged


@tagged("post_install", "-at_install")
class TestCasKardexReport(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.employee = cls.env["hr.employee"].create({"name": "Report Test Employee", "company_id": cls.env.company.id})
        attendance = cls.env["cas.attendance.day"].with_context(cas_attendance_engine=True).create({
            "employee_id": cls.employee.id, "work_date": date(2026, 7, 1), "attendance_mode": "simple", "tolerance_minutes": 5,
        })
        period = cls.env["cas.kardex.period"].period_for(cls.env.company, date(2026, 7, 1))
        cls.day = cls.env["cas.kardex.day"].with_context(cas_kardex_engine=True).create({
            "employee_id": cls.employee.id, "work_date": date(2026, 7, 1),
            "attendance_day_id": attendance.id, "period_id": period.id,
        })
        cls.day.with_context(cas_kardex_engine=True).write({
            "planned_base_minutes": 480, "presence_minutes": 510, "deducted_break_minutes": 30,
            "net_work_minutes": 480, "credited_base_minutes": 480, "absence_minutes": 0,
            "tardy_minutes": 5, "early_exit_minutes": 0, "mandatory_overtime_minutes": 0,
            "discretionary_overtime_minutes": 20, "approved_overtime_minutes": 15,
            "holiday_work_minutes": 0, "state": "final",
        })

    def test_01_excel_contains_detail_and_summary_with_minute_precision(self):
        content, count = self.env["cas.kardex.report.service"].build_xlsx(
            date(2026, 7, 1), date(2026, 7, 31), include_detail=True, include_summary=True,
        )
        self.assertEqual(count, 1)
        self.assertTrue(content.startswith(b"PK"))
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        self.assertEqual(wb.sheetnames, ["تفصیلی", "خلاصه"])
        self.assertEqual(wb["تفصیلی"].cell(2, 2).value, self.employee.name)
        self.assertEqual(int(wb["تفصیلی"].cell(2, 6).value.total_seconds() // 60), 510)
        self.assertEqual(wb["خلاصه"].cell(2, 3).value, 1)
        self.assertEqual(int(wb["خلاصه"].cell(2, 4).value.total_seconds() // 60), 480)

    def test_02_filters_can_return_empty_valid_workbook(self):
        content, count = self.env["cas.kardex.report.service"].build_xlsx(
            date(2026, 8, 1), date(2026, 8, 31), include_detail=False, include_summary=True,
        )
        self.assertEqual(count, 0)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        self.assertEqual(wb.sheetnames, ["خلاصه"])

    def test_03_wizard_rejects_invalid_range_or_empty_selection(self):
        wizard = self.env["cas.kardex.report.wizard"].create({
            "date_from": date(2026, 7, 31), "date_to": date(2026, 7, 1),
            "include_detail": True, "include_summary": False,
        })
        with self.assertRaises(ValidationError): wizard.action_export()
        wizard.write({"date_from": date(2026, 7, 1), "include_detail": False})
        with self.assertRaises(ValidationError): wizard.action_export()
